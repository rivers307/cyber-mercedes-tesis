from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, F
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
import json
import os
from django.conf import settings

from estaciones.models import Estacion, SesionPC
from sublimacion.models import Pedido, Producto  # <--- IMPORTACIÓN CORRECTA
from inventario.models import MovimientoInventario, Insumo, Auditoria, ConciliacionInventario
from usuarios.decorators import empleado_required, admin_required
from .models import Ingreso, TasaCambio, PrecioServicio
# NOTA: HistorialEstado no se usa en este archivo, por eso no se importa


# ==================== DASHBOARD PRINCIPAL ====================
@empleado_required
def dashboard_principal(request):
    from usuarios.models import Usuario
    
    hoy = timezone.now().date()
    
    ingresos_hoy = Ingreso.objects.filter(fecha__date=hoy).aggregate(total=Sum('monto'))['total'] or 0
    ingresos_mes = Ingreso.objects.filter(fecha__month=hoy.month).aggregate(total=Sum('monto'))['total'] or 0
    
    sesiones_hoy = SesionPC.objects.filter(hora_fin__date=hoy, pagado=True).count()
    ingresos_pcs_hoy = SesionPC.objects.filter(hora_fin__date=hoy, pagado=True).aggregate(total=Sum('monto_cobrado'))['total'] or 0
    
    abonos_hoy = Ingreso.objects.filter(fecha__date=hoy, tipo='sublimacion_abono').aggregate(total=Sum('monto'))['total'] or 0
    
    pedidos_pendientes = Pedido.objects.filter(estado='pendiente').count()
    ultimos_pedidos = Pedido.objects.all().order_by('-fecha_pedido')[:5]
    
    from estaciones.models import Estacion
    estaciones = Estacion.objects.all()
    total_estaciones = estaciones.count()
    estaciones_libres = estaciones.filter(estado='libre').count()
    estaciones_ocupadas = estaciones.filter(estado='ocupada').count()
    estaciones_mantenimiento = estaciones.filter(estado='mantenimiento').count()
    
    usuarios_activos = Usuario.objects.count()
    
    context = {
        'ingresos_hoy': ingresos_hoy,
        'ingresos_mes': ingresos_mes,
        'ingresos_pcs_hoy': ingresos_pcs_hoy,
        'sesiones_hoy': sesiones_hoy,
        'abonos_hoy': abonos_hoy,
        'total_ingresos_hoy': float(ingresos_hoy) + float(ingresos_pcs_hoy),
        'total_estaciones': total_estaciones,
        'estaciones_libres': estaciones_libres,
        'estaciones_ocupadas': estaciones_ocupadas,
        'estaciones_mantenimiento': estaciones_mantenimiento,
        'pedidos_pendientes': pedidos_pendientes,
        'ultimos_pedidos': ultimos_pedidos,
        'usuarios_activos': usuarios_activos,
        'usuario': request.user,
    }
    return render(request, 'reportes/dashboard_principal.html', context)


# ==================== REPORTE DE VENTAS ====================
@empleado_required
def ventas_view(request):
    """Reporte de ventas con gráficos"""
    hoy_local = timezone.localtime(timezone.now()).date()
    inicio_semana = hoy_local - timedelta(days=hoy_local.weekday())
    inicio_mes = hoy_local.replace(day=1)

    # ===== PEDIDOS DE SUBLIMACIÓN =====
    ingresos_hoy_pedidos = Pedido.objects.filter(
        fecha_pedido__date=hoy_local
    ).aggregate(total=Sum('precio_total'))['total'] or 0

    ingresos_semana_pedidos = Pedido.objects.filter(
        fecha_pedido__date__gte=inicio_semana,
        fecha_pedido__date__lte=hoy_local
    ).aggregate(total=Sum('precio_total'))['total'] or 0

    ingresos_mes_pedidos = Pedido.objects.filter(
        fecha_pedido__date__gte=inicio_mes,
        fecha_pedido__date__lte=hoy_local
    ).aggregate(total=Sum('precio_total'))['total'] or 0

    # ===== INGRESOS DE PCS (estacion) =====
    ingresos_hoy_pcs = Ingreso.objects.filter(
        fecha__date=hoy_local,
        tipo='estacion'
    ).aggregate(total=Sum('monto'))['total'] or 0

    ingresos_semana_pcs = Ingreso.objects.filter(
        fecha__date__gte=inicio_semana,
        fecha__date__lte=hoy_local,
        tipo='estacion'
    ).aggregate(total=Sum('monto'))['total'] or 0

    ingresos_mes_pcs = Ingreso.objects.filter(
        fecha__date__gte=inicio_mes,
        fecha__date__lte=hoy_local,
        tipo='estacion'
    ).aggregate(total=Sum('monto'))['total'] or 0

    # ===== INGRESOS POR VENTA DE INSUMOS =====
    ingresos_hoy_insumos = Ingreso.objects.filter(
        fecha__date=hoy_local,
        tipo='insumo'
    ).aggregate(total=Sum('monto'))['total'] or 0

    ingresos_semana_insumos = Ingreso.objects.filter(
        fecha__date__gte=inicio_semana,
        fecha__date__lte=hoy_local,
        tipo='insumo'
    ).aggregate(total=Sum('monto'))['total'] or 0

    ingresos_mes_insumos = Ingreso.objects.filter(
        fecha__date__gte=inicio_mes,
        fecha__date__lte=hoy_local,
        tipo='insumo'
    ).aggregate(total=Sum('monto'))['total'] or 0

    # ===== ABONOS PENDIENTES =====
    abonos_pendientes = Pedido.objects.filter(abono__lt=F('precio_total')).aggregate(
        total=Sum(F('precio_total') - F('abono'))
    )['total'] or 0

    # ===== TOTALES GENERALES =====
    ingresos_hoy = ingresos_hoy_pedidos + ingresos_hoy_pcs + ingresos_hoy_insumos
    ingresos_semana = ingresos_semana_pedidos + ingresos_semana_pcs + ingresos_semana_insumos
    ingresos_mes = ingresos_mes_pedidos + ingresos_mes_pcs + ingresos_mes_insumos

    ultimos_pedidos = Pedido.objects.all().order_by('-fecha_pedido')[:20]

    return render(request, 'reportes/ventas.html', {
        'ingresos_hoy': ingresos_hoy,
        'ingresos_semana': ingresos_semana,
        'ingresos_mes': ingresos_mes,
        'abonos_pendientes': abonos_pendientes,
        'ultimos_pedidos': ultimos_pedidos,
    })


# ==================== PÁGINA DE REGISTRO DE VENTAS ====================
@empleado_required
def registrar_venta_page(request):
    """Página independiente para registrar ventas"""
    tasa = TasaCambio.objects.first()
    tasa_valor = float(tasa.tasa) if tasa else 60.0

    insumos = Insumo.objects.filter(stock_actual__gt=0)
    insumos_con_precios = []
    for insumo in insumos:
        precio_usd = float(getattr(insumo, 'precio_usd', 0))
        if precio_usd == 0:
            precio_usd = float(insumo.precio_unitario) / tasa_valor if tasa_valor else 0
        precio_bs = precio_usd * tasa_valor
        insumos_con_precios.append({
            'id': insumo.id,
            'nombre': insumo.nombre,
            'precio_usd': precio_usd,
            'precio_bs': precio_bs,
            'stock': insumo.stock_actual,
        })

    return render(request, 'reportes/registrar_venta.html', {
        'insumos': insumos_con_precios,
        'tasa_actual': tasa,
    })


# ==================== REGISTRAR VENTA (PROCESAR POST) ====================
@empleado_required
@login_required
def registrar_venta(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        items_json = request.POST.get('items')
        if not items_json:
            return JsonResponse({'success': False, 'error': 'No hay items en la venta'})
        
        items = json.loads(items_json)
        metodo_pago = request.POST.get('metodo_pago', 'efectivo')
        cliente_nombre = request.POST.get('cliente_nombre', '')
        descripcion = request.POST.get('descripcion', '')
        
        total_bs = Decimal('0')
        total_usd = Decimal('0')
        
        for item_data in items:
            insumo_id = item_data.get('insumo_id')
            cantidad = int(item_data.get('cantidad', 1))
            precio_usd = Decimal(str(item_data.get('precio_usd', 0)))
            
            insumo = get_object_or_404(Insumo, id=insumo_id)
            
            if insumo.stock_actual < cantidad:
                return JsonResponse({
                    'success': False,
                    'error': f'Stock insuficiente de {insumo.nombre}'
                })
            
            tasa = TasaCambio.objects.first()
            tasa_valor = Decimal(str(tasa.tasa)) if tasa else Decimal('60')
            precio_bs = precio_usd * tasa_valor
            subtotal_bs = precio_bs * cantidad
            subtotal_usd = precio_usd * cantidad
            
            total_bs += subtotal_bs
            total_usd += subtotal_usd
            
            MovimientoInventario.objects.create(
                insumo=insumo,
                tipo='salida',
                cantidad=cantidad,
                stock_anterior=insumo.stock_actual,
                stock_nuevo=insumo.stock_actual - cantidad,
                motivo=f"Venta al {'cliente' if cliente_nombre else 'mostrador'}",
                realizado_por=request.user
            )
            insumo.stock_actual -= cantidad
            insumo.save()
        
        Ingreso.objects.create(
            tipo='insumo',
            monto=total_bs,
            metodo_pago=metodo_pago,
            descripcion=descripcion or f"Venta de {len(items)} items",
            registrado_por=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Venta registrada: {len(items)} items - Bs {total_bs:.2f} (${total_usd:.2f})'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ==================== TABULADOR DE PRECIOS ====================
@empleado_required
def tabulador_precios(request):
    """Vista del tabulador de precios"""
    tasa = TasaCambio.objects.first()
    precios = PrecioServicio.objects.filter(activo=True)
    
    return render(request, 'reportes/tabulador_precios.html', {
        'tasa': tasa,
        'precios': precios,
    })


@admin_required
def actualizar_tasa(request):
    """Actualizar la tasa de cambio (manual)"""
    if request.method == 'POST':
        tasa_valor = request.POST.get('tasa')
        fuente = request.POST.get('fuente', 'Manual')
        
        try:
            tasa = Decimal(tasa_valor)
            TasaCambio.objects.create(
                tasa=tasa,
                actualizada_por=request.user,
                fuente=fuente
            )
            messages.success(request, f'✅ Tasa actualizada: Bs {tasa:.2f} / USD')
        except:
            messages.error(request, '❌ Valor inválido')
        
        return redirect('tabulador_precios')
    
    return redirect('tabulador_precios')


@admin_required
def actualizar_tasa_api(request):
    """Actualizar tasa automáticamente desde API externa"""
    try:
        import requests
        response = requests.get('https://api.exchangerate-api.com/v4/latest/USD')
        data = response.json()
        tasa_valor = data.get('rates', {}).get('VES')
        
        if tasa_valor:
            tasa = Decimal(str(tasa_valor))
            TasaCambio.objects.create(
                tasa=tasa,
                actualizada_por=request.user,
                fuente='API Exchangerate'
            )
            messages.success(request, f'✅ Tasa actualizada desde API: Bs {tasa:.2f} / USD')
        else:
            messages.error(request, '❌ No se pudo obtener la tasa desde la API')
            
    except Exception as e:
        messages.error(request, f'❌ Error al obtener tasa: {str(e)}')
    
    return redirect('tabulador_precios')


@admin_required
def editar_precio(request):
    """Editar el precio de un servicio"""
    if request.method == 'POST':
        servicio_id = request.POST.get('servicio_id')
        precio_usd = request.POST.get('precio_usd')
        nombre_mostrar = request.POST.get('nombre_mostrar')
        
        try:
            precio = get_object_or_404(PrecioServicio, id=servicio_id)
            precio.precio_usd = Decimal(precio_usd)
            if nombre_mostrar:
                precio.nombre_mostrar = nombre_mostrar
            precio.save()
            messages.success(request, f'✅ Precio actualizado: {precio.nombre_mostrar}')
        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')
        
        return redirect('tabulador_precios')
    
    return redirect('tabulador_precios')


# ==================== EXPORTAR REPORTE DE VENTAS A EXCEL ====================
@empleado_required
def exportar_ventas_excel(request):
    """Exportar reporte de ventas a Excel con formato profesional"""
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_cyber_comunal.png')
    try:
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = 120
            img.height = 80
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 90
        else:
            ws.merge_cells('A1:G1')
            ws['A1'].value = "CYBER COMUNAL"
            ws['A1'].font = Font(name='Arial', size=20, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
    except:
        ws.merge_cells('A1:G1')
        ws['A1'].value = "CYBER COMUNAL"
        ws['A1'].font = Font(name='Arial', size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'].value = "TECH & SUBLIMACIÓN"
    ws['A2'].font = Font(name='Arial', size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:G3')
    ws['A3'].value = "Dirección: Av. Principal, C.C. Las Mercedes, Nivel 1, Local 4B"
    ws['A3'].font = Font(name='Arial', size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A4:G4')
    ws['A4'].value = "Teléfono: +58 412-5551234  |  Email: cyber.comunal.ve@gmail.com"
    ws['A4'].font = Font(name='Arial', size=9)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A5:G5')
    ws['A5'].border = Border(bottom=Side(style='medium', color='7C3AED'))
    
    ws.merge_cells('A6:G6')
    ws['A6'].value = "REPORTE DE VENTAS"
    ws['A6'].font = Font(name='Arial', size=16, bold=True, color='7C3AED')
    ws['A6'].alignment = Alignment(horizontal='center')
    
    hoy = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells('A7:G7')
    ws['A7'].value = f"Fecha de Emisión: {hoy}"
    ws['A7'].font = Font(name='Arial', size=10)
    ws['A7'].alignment = Alignment(horizontal='center')
    
    ws.row_dimensions[8].height = 20
    
    headers = ['ID', 'Cliente', 'Producto', 'Cantidad', 'Precio Unit.', 'Total (Bs)', 'Estado']
    col_widths = [10, 25, 30, 12, 18, 18, 20]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=i, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    
    pedidos = Pedido.objects.all().order_by('-fecha_pedido')
    row_num = 10
    total_general = 0
    
    for pedido in pedidos:
        ws.cell(row=row_num, column=1, value=pedido.id)
        ws.cell(row=row_num, column=2, value=pedido.nombre_cliente)
        ws.cell(row=row_num, column=3, value=pedido.producto.nombre)
        ws.cell(row=row_num, column=4, value=pedido.cantidad)
        ws.cell(row=row_num, column=5, value=float(pedido.producto.precio_base))
        ws.cell(row=row_num, column=6, value=float(pedido.precio_total))
        ws.cell(row=row_num, column=7, value=pedido.get_estado_display())
        
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=5).number_format = '#,##0.00'
        ws.cell(row=row_num, column=6).number_format = '#,##0.00'
        
        total_general += float(pedido.precio_total)
        row_num += 1
    
    ws.merge_cells(f'A{row_num}:E{row_num}')
    ws.cell(row=row_num, column=1, value="TOTAL GENERAL").font = Font(bold=True)
    ws.cell(row=row_num, column=6, value=total_general).number_format = '#,##0.00'
    ws.cell(row=row_num, column=6).font = Font(bold=True)
    
    row_num += 3
    ws.merge_cells(f'A{row_num}:B{row_num}')
    ws.cell(row=row_num, column=1, value="FIRMA AUDITOR RESPONSABLE").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'C{row_num}:D{row_num}')
    ws.cell(row=row_num, column=3, value="FIRMA COMISIÓN DE CONTRALORÍA").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'E{row_num}:G{row_num}')
    ws.cell(row=row_num, column=5, value="FIRMA DE RECEPCIÓN").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
    
    row_num += 1
    for col in [1, 3, 5]:
        ws.merge_cells(f'{get_column_letter(col)}{row_num}:{get_column_letter(col+1)}{row_num}')
        ws.cell(row=row_num, column=col, value="_________________________").alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="(admin / AD01)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=3, value="(Punto de Control)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=5, value="(____ / ____ / ____)").alignment = Alignment(horizontal='center')
    
    row_num += 3
    ws.merge_cells(f'A{row_num}:G{row_num}')
    ws.cell(row=row_num, column=1, value="Documento generado por el Sistema Inteligente de Gestión del Cyber Comunitario 'Ntra. Sra. de Mercedes'")
    ws.cell(row=row_num, column=1).font = Font(size=9, italic=True, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:G{row_num}')
    ws.cell(row=row_num, column=1, value=f"Generado el: {hoy}")
    ws.cell(row=row_num, column=1).font = Font(size=9, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_ventas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response


# ==================== EXPORTAR REPORTE DE ESTACIONES A EXCEL ====================
@empleado_required
def exportar_estaciones_excel(request):
    """Exportar reporte de estaciones a Excel con formato profesional"""
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from decimal import Decimal
    
    tasa = TasaCambio.objects.first()
    tasa_valor = float(tasa.tasa) if tasa else 60.0
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Estaciones"
    
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_cyber_comunal.png')
    try:
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = 120
            img.height = 80
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 90
        else:
            ws.merge_cells('A1:G1')
            ws['A1'].value = "CYBER COMUNAL"
            ws['A1'].font = Font(name='Arial', size=20, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
    except:
        ws.merge_cells('A1:G1')
        ws['A1'].value = "CYBER COMUNAL"
        ws['A1'].font = Font(name='Arial', size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'].value = "TECH & SUBLIMACIÓN"
    ws['A2'].font = Font(name='Arial', size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:G3')
    ws['A3'].value = "Dirección: Av. Principal, C.C. Las Mercedes, Nivel 1, Local 4B"
    ws['A3'].font = Font(name='Arial', size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A4:G4')
    ws['A4'].value = "Teléfono: +58 412-5551234  |  Email: cyber.comunal.ve@gmail.com"
    ws['A4'].font = Font(name='Arial', size=9)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A5:G5')
    ws['A5'].border = Border(bottom=Side(style='medium', color='7C3AED'))
    
    ws.merge_cells('A6:G6')
    ws['A6'].value = "REPORTE DE ESTACIONES"
    ws['A6'].font = Font(name='Arial', size=16, bold=True, color='7C3AED')
    ws['A6'].alignment = Alignment(horizontal='center')
    
    hoy = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells('A7:G7')
    ws['A7'].value = f"Fecha de Emisión: {hoy}"
    ws['A7'].font = Font(name='Arial', size=10)
    ws['A7'].alignment = Alignment(horizontal='center')
    
    ws.row_dimensions[8].height = 20
    
    # Encabezados de tabla
    headers = ['Número', 'Estado', 'Precio Hora (Bs)', 'Precio Hora (USD)', 'Último Uso', 'Tiempo Acumulado (min)']
    col_widths = [12, 20, 18, 18, 25, 22]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=i, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    
    # Datos
    estaciones = Estacion.objects.all().order_by('numero')
    row_num = 10
    total_bs = 0
    total_usd = 0
    for estacion in estaciones:
        precio_bs = float(estacion.precio_hora)
        precio_usd = precio_bs / tasa_valor if tasa_valor > 0 else 0
        total_bs += precio_bs
        total_usd += precio_usd
        
        ws.cell(row=row_num, column=1, value=estacion.numero)
        ws.cell(row=row_num, column=2, value=estacion.get_estado_display())
        ws.cell(row=row_num, column=3, value=precio_bs)
        ws.cell(row=row_num, column=4, value=precio_usd)
        ultimo = "Nunca"
        if hasattr(estacion, 'ultimo_uso') and estacion.ultimo_uso:
            ultimo = estacion.ultimo_uso.strftime('%d/%m/%Y %H:%M')
        elif hasattr(estacion, 'hora_inicio') and estacion.hora_inicio:
            ultimo = estacion.hora_inicio.strftime('%d/%m/%Y %H:%M')
        ws.cell(row=row_num, column=5, value=ultimo)
        ws.cell(row=row_num, column=6, value=estacion.tiempo_acumulado if hasattr(estacion, 'tiempo_acumulado') else 0)
        
        ws.cell(row=row_num, column=3).number_format = '#,##0.00'
        ws.cell(row=row_num, column=4).number_format = '#,##0.00'
        row_num += 1
    
    # Total de precios
    if row_num > 10:
        ws.merge_cells(f'A{row_num}:B{row_num}')
        ws.cell(row=row_num, column=1, value="TOTAL")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=3, value=total_bs).number_format = '#,##0.00'
        ws.cell(row=row_num, column=4, value=total_usd).number_format = '#,##0.00'
        row_num += 1
    
    # Firmas
    row_num += 2
    ws.merge_cells(f'A{row_num}:B{row_num}')
    ws.cell(row=row_num, column=1, value="FIRMA AUDITOR RESPONSABLE").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'C{row_num}:D{row_num}')
    ws.cell(row=row_num, column=3, value="FIRMA COMISIÓN DE CONTRALORÍA").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'E{row_num}:G{row_num}')
    ws.cell(row=row_num, column=5, value="FIRMA DE RECEPCIÓN").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
    
    row_num += 1
    for col in [1, 3, 5]:
        ws.merge_cells(f'{get_column_letter(col)}{row_num}:{get_column_letter(col+1)}{row_num}')
        ws.cell(row=row_num, column=col, value="_________________________").alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="(admin / AD01)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=3, value="(Punto de Control)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=5, value="(____ / ____ / ____)").alignment = Alignment(horizontal='center')
    
    row_num += 2
    ws.merge_cells(f'A{row_num}:G{row_num}')
    ws.cell(row=row_num, column=1, value="Documento generado por el Sistema Inteligente de Gestión del Cyber Comunitario 'Ntra. Sra. de Mercedes'")
    ws.cell(row=row_num, column=1).font = Font(size=9, italic=True, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:G{row_num}')
    ws.cell(row=row_num, column=1, value=f"Generado el: {hoy}")
    ws.cell(row=row_num, column=1).font = Font(size=9, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_estaciones_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response


# ==================== EXPORTAR REPORTE DE SUBLIMACIÓN A EXCEL ====================
@empleado_required
def exportar_sublimacion_excel(request):
    """Exportar reporte de sublimación a Excel con formato profesional"""
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from decimal import Decimal
    
    tasa = TasaCambio.objects.first()
    tasa_valor = float(tasa.tasa) if tasa else 60.0
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Sublimación"
    
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_cyber_comunal.png')
    try:
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = 120
            img.height = 80
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 90
        else:
            ws.merge_cells('A1:H1')
            ws['A1'].value = "CYBER COMUNAL"
            ws['A1'].font = Font(name='Arial', size=20, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
    except:
        ws.merge_cells('A1:H1')
        ws['A1'].value = "CYBER COMUNAL"
        ws['A1'].font = Font(name='Arial', size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'].value = "TECH & SUBLIMACIÓN"
    ws['A2'].font = Font(name='Arial', size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:H3')
    ws['A3'].value = "Dirección: Av. Principal, C.C. Las Mercedes, Nivel 1, Local 4B"
    ws['A3'].font = Font(name='Arial', size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A4:H4')
    ws['A4'].value = "Teléfono: +58 412-5551234  |  Email: cyber.comunal.ve@gmail.com"
    ws['A4'].font = Font(name='Arial', size=9)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A5:H5')
    ws['A5'].border = Border(bottom=Side(style='medium', color='7C3AED'))
    
    ws.merge_cells('A6:H6')
    ws['A6'].value = "REPORTE DE SUBLIMACIÓN"
    ws['A6'].font = Font(name='Arial', size=16, bold=True, color='7C3AED')
    ws['A6'].alignment = Alignment(horizontal='center')
    
    hoy = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells('A7:H7')
    ws['A7'].value = f"Fecha de Emisión: {hoy}"
    ws['A7'].font = Font(name='Arial', size=10)
    ws['A7'].alignment = Alignment(horizontal='center')
    
    ws.row_dimensions[8].height = 20
    
    headers = ['ID', 'Cliente', 'Producto', 'Cantidad', 'Total (Bs)', 'Total (USD)', 'Abono (Bs)', 'Saldo (Bs)', 'Estado', 'Fecha']
    col_widths = [10, 25, 25, 12, 18, 18, 18, 18, 20, 18]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=i, value=header)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    
    pedidos = Pedido.objects.all().order_by('-fecha_pedido')
    row_num = 10
    total_bs = 0
    total_usd = 0
    total_abonos = 0
    
    for pedido in pedidos:
        precio_bs = float(pedido.precio_total)
        precio_usd = precio_bs / tasa_valor if tasa_valor > 0 else 0
        saldo = float(pedido.saldo_pendiente())
        total_bs += precio_bs
        total_usd += precio_usd
        total_abonos += float(pedido.abono)
        
        ws.cell(row=row_num, column=1, value=pedido.id)
        ws.cell(row=row_num, column=2, value=pedido.nombre_cliente)
        ws.cell(row=row_num, column=3, value=pedido.producto.nombre)
        ws.cell(row=row_num, column=4, value=pedido.cantidad)
        ws.cell(row=row_num, column=5, value=precio_bs)
        ws.cell(row=row_num, column=6, value=precio_usd)
        ws.cell(row=row_num, column=7, value=float(pedido.abono))
        ws.cell(row=row_num, column=8, value=saldo)
        ws.cell(row=row_num, column=9, value=pedido.get_estado_display())
        ws.cell(row=row_num, column=10, value=pedido.fecha_pedido.strftime('%d/%m/%Y %H:%M'))
        
        ws.cell(row=row_num, column=5).number_format = '#,##0.00'
        ws.cell(row=row_num, column=6).number_format = '#,##0.00'
        ws.cell(row=row_num, column=7).number_format = '#,##0.00'
        ws.cell(row=row_num, column=8).number_format = '#,##0.00'
        row_num += 1
    
    if row_num > 10:
        ws.merge_cells(f'A{row_num}:D{row_num}')
        ws.cell(row=row_num, column=1, value="TOTALES")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=5, value=total_bs).number_format = '#,##0.00'
        ws.cell(row=row_num, column=6, value=total_usd).number_format = '#,##0.00'
        ws.cell(row=row_num, column=7, value=total_abonos).number_format = '#,##0.00'
        row_num += 1
    
    row_num += 2
    ws.merge_cells(f'A{row_num}:B{row_num}')
    ws.cell(row=row_num, column=1, value="FIRMA AUDITOR RESPONSABLE").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'C{row_num}:D{row_num}')
    ws.cell(row=row_num, column=3, value="FIRMA COMISIÓN DE CONTRALORÍA").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'E{row_num}:H{row_num}')
    ws.cell(row=row_num, column=5, value="FIRMA DE RECEPCIÓN").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
    
    row_num += 1
    for col in [1, 3, 5]:
        ws.merge_cells(f'{get_column_letter(col)}{row_num}:{get_column_letter(col+1)}{row_num}')
        ws.cell(row=row_num, column=col, value="_________________________").alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="(admin / AD01)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=3, value="(Punto de Control)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=5, value="(____ / ____ / ____)").alignment = Alignment(horizontal='center')
    
    row_num += 2
    ws.merge_cells(f'A{row_num}:H{row_num}')
    ws.cell(row=row_num, column=1, value="Documento generado por el Sistema Inteligente de Gestión del Cyber Comunitario 'Ntra. Sra. de Mercedes'")
    ws.cell(row=row_num, column=1).font = Font(size=9, italic=True, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:H{row_num}')
    ws.cell(row=row_num, column=1, value=f"Generado el: {hoy}")
    ws.cell(row=row_num, column=1).font = Font(size=9, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_sublimacion_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response


# ==================== EXPORTAR REPORTE DE INVENTARIO A EXCEL ====================
@empleado_required
def exportar_inventario_excel(request):
    """Exportar reporte de inventario a Excel con formato profesional"""
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from decimal import Decimal
    
    tasa = TasaCambio.objects.first()
    tasa_valor = float(tasa.tasa) if tasa else 60.0
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Inventario"
    
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_cyber_comunal.png')
    try:
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = 120
            img.height = 80
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 90
        else:
            ws.merge_cells('A1:H1')
            ws['A1'].value = "CYBER COMUNAL"
            ws['A1'].font = Font(name='Arial', size=20, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
    except:
        ws.merge_cells('A1:H1')
        ws['A1'].value = "CYBER COMUNAL"
        ws['A1'].font = Font(name='Arial', size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'].value = "TECH & SUBLIMACIÓN"
    ws['A2'].font = Font(name='Arial', size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:H3')
    ws['A3'].value = "Dirección: Av. Principal, C.C. Las Mercedes, Nivel 1, Local 4B"
    ws['A3'].font = Font(name='Arial', size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A4:H4')
    ws['A4'].value = "Teléfono: +58 412-5551234  |  Email: cyber.comunal.ve@gmail.com"
    ws['A4'].font = Font(name='Arial', size=9)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A5:H5')
    ws['A5'].border = Border(bottom=Side(style='medium', color='7C3AED'))
    
    ws.merge_cells('A6:H6')
    ws['A6'].value = "REPORTE DE INVENTARIO"
    ws['A6'].font = Font(name='Arial', size=16, bold=True, color='7C3AED')
    ws['A6'].alignment = Alignment(horizontal='center')
    
    hoy = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells('A7:H7')
    ws['A7'].value = f"Fecha de Emisión: {hoy}"
    ws['A7'].font = Font(name='Arial', size=10)
    ws['A7'].alignment = Alignment(horizontal='center')
    
    ws.row_dimensions[8].height = 20
    
    headers = ['Código', 'Nombre', 'Categoría', 'Stock', 'Stock Mínimo', 'Precio (Bs)', 'Precio (USD)', 'Valor Total (Bs)']
    col_widths = [15, 25, 20, 12, 12, 18, 18, 20]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=i, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    
    insumos = Insumo.objects.all().order_by('nombre')
    row_num = 10
    total_valor_bs = 0
    for insumo in insumos:
        valor_bs = insumo.stock_actual * float(insumo.precio_unitario)
        total_valor_bs += valor_bs
        precio_usd = float(insumo.precio_usd) if hasattr(insumo, 'precio_usd') else 0
        precio_bs = float(insumo.precio_unitario)
        
        ws.cell(row=row_num, column=1, value=insumo.codigo)
        ws.cell(row=row_num, column=2, value=insumo.nombre)
        ws.cell(row=row_num, column=3, value=insumo.categoria.nombre if insumo.categoria else 'Sin categoría')
        ws.cell(row=row_num, column=4, value=insumo.stock_actual)
        ws.cell(row=row_num, column=5, value=insumo.stock_minimo)
        ws.cell(row=row_num, column=6, value=precio_bs)
        ws.cell(row=row_num, column=7, value=precio_usd)
        ws.cell(row=row_num, column=8, value=valor_bs)
        
        ws.cell(row=row_num, column=6).number_format = '#,##0.00'
        ws.cell(row=row_num, column=7).number_format = '#,##0.00'
        ws.cell(row=row_num, column=8).number_format = '#,##0.00'
        row_num += 1
    
    if row_num > 10:
        ws.merge_cells(f'A{row_num}:G{row_num}')
        ws.cell(row=row_num, column=1, value="TOTAL")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=8, value=total_valor_bs).number_format = '#,##0.00'
        row_num += 1
    
    row_num += 2
    ws.merge_cells(f'A{row_num}:B{row_num}')
    ws.cell(row=row_num, column=1, value="FIRMA AUDITOR RESPONSABLE").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'C{row_num}:D{row_num}')
    ws.cell(row=row_num, column=3, value="FIRMA COMISIÓN DE CONTRALORÍA").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'E{row_num}:H{row_num}')
    ws.cell(row=row_num, column=5, value="FIRMA DE RECEPCIÓN").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
    
    row_num += 1
    for col in [1, 3, 5]:
        ws.merge_cells(f'{get_column_letter(col)}{row_num}:{get_column_letter(col+1)}{row_num}')
        ws.cell(row=row_num, column=col, value="_________________________").alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="(admin / AD01)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=3, value="(Punto de Control)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=5, value="(____ / ____ / ____)").alignment = Alignment(horizontal='center')
    
    row_num += 2
    ws.merge_cells(f'A{row_num}:H{row_num}')
    ws.cell(row=row_num, column=1, value="Documento generado por el Sistema Inteligente de Gestión del Cyber Comunitario 'Ntra. Sra. de Mercedes'")
    ws.cell(row=row_num, column=1).font = Font(size=9, italic=True, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:H{row_num}')
    ws.cell(row=row_num, column=1, value=f"Generado el: {hoy}")
    ws.cell(row=row_num, column=1).font = Font(size=9, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_inventario_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response


# ==================== EXPORTAR AUDITORÍA DE VENTAS A EXCEL ====================
@admin_required
def exportar_auditoria_ventas_excel(request):
    """Exportar auditoría de ventas a Excel con formato profesional"""
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from decimal import Decimal
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría de Ventas"
    
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_cyber_comunal.png')
    try:
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = 120
            img.height = 80
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 90
        else:
            ws.merge_cells('A1:G1')
            ws['A1'].value = "CYBER COMUNAL"
            ws['A1'].font = Font(name='Arial', size=20, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
    except:
        ws.merge_cells('A1:G1')
        ws['A1'].value = "CYBER COMUNAL"
        ws['A1'].font = Font(name='Arial', size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'].value = "TECH & SUBLIMACIÓN"
    ws['A2'].font = Font(name='Arial', size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:G3')
    ws['A3'].value = "Dirección: Av. Principal, C.C. Las Mercedes, Nivel 1, Local 4B"
    ws['A3'].font = Font(name='Arial', size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A4:G4')
    ws['A4'].value = "Teléfono: +58 412-5551234  |  Email: cyber.comunal.ve@gmail.com"
    ws['A4'].font = Font(name='Arial', size=9)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A5:G5')
    ws['A5'].border = Border(bottom=Side(style='medium', color='7C3AED'))
    
    ws.merge_cells('A6:G6')
    ws['A6'].value = "AUDITORÍA DE VENTAS"
    ws['A6'].font = Font(name='Arial', size=16, bold=True, color='7C3AED')
    ws['A6'].alignment = Alignment(horizontal='center')
    
    hoy = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells('A7:G7')
    ws['A7'].value = f"Fecha de Emisión: {hoy}"
    ws['A7'].font = Font(name='Arial', size=10)
    ws['A7'].alignment = Alignment(horizontal='center')
    
    ws.row_dimensions[8].height = 20
    
    headers = ['Fecha', 'Tipo', 'Descripción', 'Monto (Bs)', 'Método de Pago', 'Registrado Por']
    col_widths = [20, 25, 35, 18, 20, 18]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=i, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    
    transacciones = Ingreso.objects.all().order_by('-fecha')
    row_num = 10
    total_ventas = 0
    for t in transacciones:
        monto = float(t.monto)
        total_ventas += monto
        
        ws.cell(row=row_num, column=1, value=t.fecha.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=2, value=t.get_tipo_display())
        ws.cell(row=row_num, column=3, value=t.descripcion)
        ws.cell(row=row_num, column=4, value=monto)
        ws.cell(row=row_num, column=5, value=t.get_metodo_pago_display() if t.metodo_pago else 'Efectivo')
        ws.cell(row=row_num, column=6, value=t.registrado_por.username)
        
        ws.cell(row=row_num, column=4).number_format = '#,##0.00'
        row_num += 1
    
    if row_num > 10:
        ws.merge_cells(f'A{row_num}:C{row_num}')
        ws.cell(row=row_num, column=1, value="TOTAL")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=4, value=total_ventas).number_format = '#,##0.00'
        row_num += 1
    
    row_num += 2
    ws.merge_cells(f'A{row_num}:B{row_num}')
    ws.cell(row=row_num, column=1, value="FIRMA AUDITOR RESPONSABLE").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'C{row_num}:D{row_num}')
    ws.cell(row=row_num, column=3, value="FIRMA COMISIÓN DE CONTRALORÍA").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'E{row_num}:G{row_num}')
    ws.cell(row=row_num, column=5, value="FIRMA DE RECEPCIÓN").font = Font(size=10, bold=True)
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
    
    row_num += 1
    for col in [1, 3, 5]:
        ws.merge_cells(f'{get_column_letter(col)}{row_num}:{get_column_letter(col+1)}{row_num}')
        ws.cell(row=row_num, column=col, value="_________________________").alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="(admin / AD01)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=3, value="(Punto de Control)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=5, value="(____ / ____ / ____)").alignment = Alignment(horizontal='center')
    
    row_num += 2
    ws.merge_cells(f'A{row_num}:G{row_num}')
    ws.cell(row=row_num, column=1, value="Documento generado por el Sistema Inteligente de Gestión del Cyber Comunitario 'Ntra. Sra. de Mercedes'")
    ws.cell(row=row_num, column=1).font = Font(size=9, italic=True, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:G{row_num}')
    ws.cell(row=row_num, column=1, value=f"Generado el: {hoy}")
    ws.cell(row=row_num, column=1).font = Font(size=9, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=auditoria_ventas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response


# ==================== API PARA VENTAS (GRÁFICOS) ====================
@empleado_required
def api_ventas_data(request):
    periodo = request.GET.get('periodo', 'diario')
    hoy = timezone.now().date()
    
    if periodo == 'diario':
        fechas = [hoy - timedelta(days=i) for i in range(6, -1, -1)]
        ventas = []
        for fecha in fechas:
            total_pedidos = Pedido.objects.filter(
                fecha_pedido__date=fecha
            ).aggregate(total=Sum('precio_total'))['total'] or 0
            total_pcs = Ingreso.objects.filter(
                fecha__date=fecha,
                tipo='estacion'
            ).aggregate(total=Sum('monto'))['total'] or 0
            total_insumos = Ingreso.objects.filter(
                fecha__date=fecha,
                tipo='insumo'
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            ventas.append({
                'fecha': fecha.strftime('%d/%m'),
                'total': float(total_pedidos) + float(total_pcs) + float(total_insumos)
            })
        return JsonResponse({'ventas': ventas})
    
    elif periodo == 'semanal':
        ventas = []
        for i in range(11, -1, -1):
            inicio_semana = hoy - timedelta(weeks=i)
            fin_semana = inicio_semana + timedelta(days=6)
            total_pedidos = Pedido.objects.filter(
                fecha_pedido__date__gte=inicio_semana,
                fecha_pedido__date__lte=fin_semana
            ).aggregate(total=Sum('precio_total'))['total'] or 0
            total_pcs = Ingreso.objects.filter(
                fecha__date__gte=inicio_semana,
                fecha__date__lte=fin_semana,
                tipo='estacion'
            ).aggregate(total=Sum('monto'))['total'] or 0
            total_insumos = Ingreso.objects.filter(
                fecha__date__gte=inicio_semana,
                fecha__date__lte=fin_semana,
                tipo='insumo'
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            ventas.append(float(total_pedidos) + float(total_pcs) + float(total_insumos))
        return JsonResponse({'ventas': ventas})
    
    else:
        meses = []
        ventas = []
        for i in range(5, -1, -1):
            fecha = hoy.replace(day=1) - timedelta(days=30*i)
            total_pedidos = Pedido.objects.filter(
                fecha_pedido__year=fecha.year,
                fecha_pedido__month=fecha.month
            ).aggregate(total=Sum('precio_total'))['total'] or 0
            total_pcs = Ingreso.objects.filter(
                fecha__year=fecha.year,
                fecha__month=fecha.month,
                tipo='estacion'
            ).aggregate(total=Sum('monto'))['total'] or 0
            total_insumos = Ingreso.objects.filter(
                fecha__year=fecha.year,
                fecha__month=fecha.month,
                tipo='insumo'
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            meses.append(fecha.strftime('%B'))
            ventas.append(float(total_pedidos) + float(total_pcs) + float(total_insumos))
        return JsonResponse({'ventas': ventas, 'meses': meses})


# ==================== REPORTE DE ESTACIONES ====================
@empleado_required
def estaciones_view(request):
    """Reporte de estaciones con precios en Bs y USD"""
    tasa = TasaCambio.objects.first()
    tasa_valor = float(tasa.tasa) if tasa else 60.0

    estaciones = Estacion.objects.all().order_by('numero')
    estaciones_data = []
    for estacion in estaciones:
        precio_usd = float(estacion.precio_hora) / tasa_valor if tasa_valor > 0 else 0
        tiempo_acumulado = estacion.tiempo_acumulado if hasattr(estacion, 'tiempo_acumulado') else 0
        ultimo_uso = None
        if hasattr(estacion, 'ultimo_uso') and estacion.ultimo_uso:
            ultimo_uso = estacion.ultimo_uso
        elif hasattr(estacion, 'hora_inicio') and estacion.hora_inicio:
            ultimo_uso = estacion.hora_inicio
        
        estaciones_data.append({
            'id': estacion.id,
            'numero': estacion.numero,
            'estado': estacion.estado,
            'estado_display': estacion.get_estado_display(),
            'precio_hora': float(estacion.precio_hora),
            'precio_usd': precio_usd,
            'ultimo_uso': ultimo_uso,
            'tiempo_acumulado': tiempo_acumulado,
            'hora_inicio': estacion.hora_inicio,
        })

    total = len(estaciones_data)
    libres = sum(1 for e in estaciones_data if e['estado'] == 'libre')
    ocupadas = sum(1 for e in estaciones_data if e['estado'] == 'ocupada')
    mantenimiento = sum(1 for e in estaciones_data if e['estado'] == 'mantenimiento')

    return render(request, 'reportes/estaciones.html', {
        'estaciones': estaciones_data,
        'total_estaciones': total,
        'libres': libres,
        'ocupadas': ocupadas,
        'mantenimiento': mantenimiento,
        'tasa_valor': tasa_valor,
        'tasa': tasa,
    })


@empleado_required
def api_estaciones_stats(request):
    """API para estadísticas de estaciones"""
    estaciones = Estacion.objects.all()
    stats = {
        'libres': estaciones.filter(estado='libre').count(),
        'ocupadas': estaciones.filter(estado='ocupada').count(),
        'mantenimiento': estaciones.filter(estado='mantenimiento').count(),
        'total': estaciones.count()
    }
    
    horas_pico = []
    for hora in range(8, 22):
        horas_pico.append({
            'hora': f"{hora}:00",
            'uso': estaciones.filter(estado='ocupada').count() * ((hora - 7) / 10)
        })
    stats['horas_pico'] = horas_pico
    
    return JsonResponse(stats)


# ==================== REPORTE DE SUBLIMACIÓN ====================
@empleado_required
def sublimacion_view(request):
    """Reporte de sublimación con estadísticas y precios en Bs y USD"""
    tasa = TasaCambio.objects.first()
    tasa_valor = float(tasa.tasa) if tasa else 60.0

    pedidos = Pedido.objects.all().order_by('-fecha_pedido')
    pedidos_data = []
    total_bs = 0
    total_usd = 0
    for pedido in pedidos:
        precio_bs = float(pedido.precio_total)
        precio_usd = precio_bs / tasa_valor if tasa_valor > 0 else 0
        saldo = float(pedido.saldo_pendiente())
        total_bs += precio_bs
        total_usd += precio_usd
        
        pedidos_data.append({
            'id': pedido.id,
            'cliente': pedido.nombre_cliente,
            'producto': pedido.producto.nombre,
            'cantidad': pedido.cantidad,
            'precio_bs': precio_bs,
            'precio_usd': precio_usd,
            'abono': float(pedido.abono),
            'saldo': saldo,
            'estado': pedido.estado,
            'estado_display': pedido.get_estado_display(),
            'fecha': pedido.fecha_pedido,
        })

    estados = {}
    for estado, label in Pedido.ESTADOS:
        count = Pedido.objects.filter(estado=estado).count()
        estados[label] = count

    productos = Producto.objects.annotate(
        total_vendidos=Sum('pedido__cantidad')
    ).filter(total_vendidos__gt=0).order_by('-total_vendidos')[:10]
    
    productos_data = []
    for p in productos:
        productos_data.append({
            'nombre': p.nombre,
            'total': p.total_vendidos or 0
        })

    return render(request, 'reportes/sublimacion.html', {
        'pedidos': pedidos_data,
        'total_pedidos': len(pedidos_data),
        'total_bs': total_bs,
        'total_usd': total_usd,
        'estados': estados,
        'productos': productos_data,
        'tasa_valor': tasa_valor,
        'tasa': tasa,
    })


# ==================== API PARA SUBLIMACIÓN ====================
@empleado_required
def api_sublimacion_stats(request):
    """API para estadísticas de sublimación"""
    estados = {}
    for estado, label in Pedido.ESTADOS:
        count = Pedido.objects.filter(estado=estado).count()
        estados[label] = count
    
    productos = Producto.objects.annotate(
        total_vendidos=Sum('pedido__cantidad')
    ).filter(total_vendidos__gt=0).order_by('-total_vendidos')[:10]
    
    productos_data = []
    for p in productos:
        productos_data.append({
            'nombre': p.nombre,
            'total': p.total_vendidos or 0
        })
    
    return JsonResponse({
        'estados': estados,
        'productos': productos_data
    })


# ==================== REPORTE DE INVENTARIO (CORREGIDO) ====================
@empleado_required
def inventario_view(request):
    """Reporte de inventario con precios en Bs y USD"""
    tasa = TasaCambio.objects.first()
    tasa_valor = float(tasa.tasa) if tasa else 60.0

    insumos = Insumo.objects.all().order_by('nombre')
    insumos_data = []
    total_valor_bs = 0
    total_valor_usd = 0
    
    for insumo in insumos:
        valor_bs = insumo.stock_actual * float(insumo.precio_unitario)
        valor_usd = valor_bs / tasa_valor if tasa_valor > 0 else 0
        precio_usd = float(insumo.precio_usd) if hasattr(insumo, 'precio_usd') else 0
        total_valor_bs += valor_bs
        total_valor_usd += valor_usd
        
        insumos_data.append({
            'id': insumo.id,
            'codigo': insumo.codigo,
            'nombre': insumo.nombre,
            'categoria': insumo.categoria.nombre if insumo.categoria else 'Sin categoría',
            'stock': insumo.stock_actual,
            'stock_minimo': insumo.stock_minimo,
            'precio_bs': float(insumo.precio_unitario),
            'precio_usd': precio_usd,
            'valor_bs': valor_bs,
            'valor_usd': valor_usd,
            'unidad': insumo.unidad,
        })

    stock_bajo = Insumo.objects.filter(stock_actual__lte=F('stock_minimo')).count()
    sin_stock = Insumo.objects.filter(stock_actual=0).count()

    return render(request, 'reportes/inventario.html', {
        'insumos': insumos_data,
        'total_insumos': len(insumos_data),
        'stock_bajo': stock_bajo,
        'sin_stock': sin_stock,
        'total_valor_bs': total_valor_bs,
        'total_valor_usd': total_valor_usd,
        'tasa_valor': tasa_valor,
        'tasa': tasa,
    })


@empleado_required
def api_inventario_stats(request):
    """API para estadísticas de inventario"""
    insumos = Insumo.objects.all()
    total_valor = sum(i.stock_actual * float(i.precio_unitario) for i in insumos)
    stock_bajo = insumos.filter(stock_actual__lte=F('stock_minimo')).count()
    sin_stock = insumos.filter(stock_actual=0).count()
    
    movimientos = MovimientoInventario.objects.all()[:20]
    movimientos_data = []
    for m in movimientos:
        movimientos_data.append({
            'fecha': m.fecha.strftime('%d/%m/%Y %H:%M'),
            'insumo': m.insumo.nombre,
            'tipo': m.get_tipo_display(),
            'cantidad': m.cantidad
        })
    
    return JsonResponse({
        'total_insumos': insumos.count(),
        'total_valor': float(total_valor),
        'stock_bajo': stock_bajo,
        'sin_stock': sin_stock,
        'movimientos': movimientos_data
    })


# ==================== CIERRE DE CAJA (CORREGIDO) ====================
@empleado_required
def cierre_caja(request):
    """Cierre de caja diario"""
    from decimal import Decimal
    from reportes.models import Ingreso
    from estaciones.models import SesionPC
    
    hoy = timezone.now().date()
    
    ingresos_dia = Ingreso.objects.filter(fecha__date=hoy)
    
    total_ventas = ingresos_dia.aggregate(total=Sum('monto'))['total'] or Decimal('0')
    total_abonos = ingresos_dia.filter(tipo='sublimacion_abono').aggregate(total=Sum('monto'))['total'] or Decimal('0')
    total_pendiente = Decimal('0')
    
    from sublimacion.models import Pedido
    entregados_hoy = Pedido.objects.filter(
        estado='entregado', 
        fecha_entrega__date=hoy
    ).count()
    
    metodos_pago = {
        'efectivo': ingresos_dia.filter(metodo_pago='efectivo').aggregate(total=Sum('monto'))['total'] or Decimal('0'),
        'transferencia': ingresos_dia.filter(metodo_pago='transferencia').aggregate(total=Sum('monto'))['total'] or Decimal('0'),
        'pago_movil': ingresos_dia.filter(metodo_pago='pago_movil').aggregate(total=Sum('monto'))['total'] or Decimal('0'),
        'tarjeta': ingresos_dia.filter(metodo_pago='tarjeta').aggregate(total=Sum('monto'))['total'] or Decimal('0'),
    }
    
    # Calcular total general
    total_general = total_ventas + total_abonos
    
    ventas_dia = ingresos_dia.order_by('-fecha')
    
    if request.method == 'POST' and 'cerrar' in request.POST:
        return render(request, 'reportes/cierre_exitoso.html', {
            'fecha': hoy,
            'total_ventas': total_ventas,
            'total_abonos': total_abonos,
            'total_pendiente': total_pendiente,
            'total_general': total_general,
        })
    
    return render(request, 'reportes/cierre_caja.html', {
        'fecha': hoy,
        'total_ventas': total_ventas,
        'total_abonos': total_abonos,
        'total_pendiente': total_pendiente,
        'total_general': total_general,
        'entregados_hoy': entregados_hoy,
        'metodos_pago': metodos_pago,
        'ventas_dia': ventas_dia,
    })


# ==================== EXPORTAR REPORTES (CSV) ====================
@empleado_required
def exportar_reporte(request, tipo):
    """Exportar reporte a CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="reporte_{tipo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    if tipo == 'ventas':
        writer.writerow(['ID', 'Cliente', 'Producto', 'Cantidad', 'Total (Bs)', 'Estado', 'Fecha'])
        for p in Pedido.objects.all():
            writer.writerow([
                p.id, 
                p.nombre_cliente, 
                p.producto.nombre, 
                p.cantidad, 
                f"{p.precio_total:.2f}", 
                p.get_estado_display(), 
                p.fecha_pedido.strftime('%d/%m/%Y')
            ])
        return response
    
    elif tipo == 'pedidos':
        writer.writerow(['ID', 'Cliente', 'Producto', 'Estado', 'Abono (Bs)', 'Saldo (Bs)'])
        for p in Pedido.objects.all():
            writer.writerow([
                p.id, 
                p.nombre_cliente, 
                p.producto.nombre, 
                p.get_estado_display(), 
                f"{p.abono:.2f}", 
                f"{p.saldo_pendiente():.2f}"
            ])
        return response
    
    elif tipo == 'inventario':
        writer.writerow(['Código', 'Nombre', 'Categoría', 'Stock Actual', 'Stock Mínimo', 'Precio Unitario (Bs)', 'Valor Total (Bs)'])
        for i in Insumo.objects.all():
            valor_total = i.stock_actual * float(i.precio_unitario)
            writer.writerow([
                i.codigo, 
                i.nombre, 
                i.categoria.nombre if i.categoria else 'Sin categoría', 
                i.stock_actual, 
                i.stock_minimo, 
                f"{float(i.precio_unitario):.2f}", 
                f"{valor_total:.2f}"
            ])
        return response
    
    elif tipo == 'insumos':
        writer.writerow(['ID', 'Nombre', 'Stock', 'Precio', 'Proveedor', 'Ubicación'])
        for i in Insumo.objects.all():
            writer.writerow([
                i.id, 
                i.nombre, 
                i.stock_actual, 
                f"{float(i.precio_unitario):.2f}", 
                i.proveedor or '', 
                i.ubicacion or ''
            ])
        return response
    
    elif tipo == 'estaciones':
        writer.writerow(['Número', 'Estado', 'Precio por Hora (Bs)', 'Último Uso', 'Tiempo Acumulado (min)'])
        for e in Estacion.objects.all():
            writer.writerow([
                e.numero, 
                e.get_estado_display(), 
                f"{float(e.precio_hora):.2f}", 
                e.ultimo_uso.strftime('%d/%m/%Y %H:%M') if hasattr(e, 'ultimo_uso') and e.ultimo_uso else 'Nunca', 
                e.tiempo_acumulado if hasattr(e, 'tiempo_acumulado') else 0
            ])
        return response
    
    else:
        from django.shortcuts import redirect
        return redirect('dashboard')


# ==================== AUDITORÍA DE VENTAS (SOLO ADMIN) ====================
@admin_required
def auditoria_ventas(request):
    """Auditoría de transacciones de ventas - SOLO ADMIN"""
    from decimal import Decimal
    from django.db.models import Sum
    
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tipo = request.GET.get('tipo')
    
    transacciones = Ingreso.objects.all()
    
    if fecha_desde:
        transacciones = transacciones.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        transacciones = transacciones.filter(fecha__date__lte=fecha_hasta)
    if tipo:
        transacciones = transacciones.filter(tipo=tipo)
    
    total_ventas = transacciones.aggregate(total=Sum('monto'))['total'] or Decimal('0')
    total_abonos = transacciones.filter(tipo='sublimacion_abono').aggregate(total=Sum('monto'))['total'] or Decimal('0')
    saldo_pendiente = total_ventas - total_abonos
    
    transacciones_data = []
    for t in transacciones:
        transacciones_data.append({
            'fecha': t.fecha.strftime('%d/%m/%Y %H:%M'),
            'tipo': t.tipo,
            'tipo_display': t.get_tipo_display(),
            'descripcion': t.descripcion,
            'monto': float(t.monto),
            'metodo_pago': t.get_metodo_pago_display() if t.metodo_pago else 'Efectivo',
            'registrado_por': t.registrado_por.username
        })
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return JsonResponse({
            'transacciones': transacciones_data,
            'total_ventas': float(total_ventas),
            'total_abonos': float(total_abonos),
            'saldo_pendiente': float(saldo_pendiente)
        })
    
    return render(request, 'reportes/auditoria_ventas.html', {
        'transacciones': transacciones,
        'total_ventas': total_ventas,
        'total_abonos': total_abonos,
        'saldo_pendiente': saldo_pendiente,
    })


# ==================== API PARA LISTA DE INSUMOS ====================
@login_required
def api_insumos_lista(request):
    """API para obtener la lista de insumos"""
    insumos = Insumo.objects.all()
    data = []
    for i in insumos:
        data.append({
            'codigo': i.codigo,
            'nombre': i.nombre,
            'categoria': i.categoria.nombre if i.categoria else 'Sin categoría',
            'stock_actual': i.stock_actual,
            'stock_minimo': i.stock_minimo,
            'precio_unitario': float(i.precio_unitario),
        })
    return JsonResponse({'insumos': data})


# ==================== PANEL DE CLIENTE ====================
@login_required
def panel_cliente(request):
    """Panel principal del cliente"""
    from sublimacion.models import Pedido
    
    pedidos = Pedido.objects.filter(nombre_cliente=request.user.username)
    
    return render(request, 'reportes/panel_cliente.html', {
        'pedidos': pedidos,
        'usuario': request.user
    })


@login_required
def detalle_pedido_cliente(request, id):
    """Ver detalle de un pedido específico del cliente"""
    from sublimacion.models import Pedido
    
    pedido = get_object_or_404(Pedido, id=id)
    
    if pedido.nombre_cliente != request.user.username:
        messages.error(request, 'No tienes permiso para ver este pedido')
        return redirect('panel_cliente')
    
    return render(request, 'reportes/detalle_pedido_cliente.html', {
        'pedido': pedido,
        'historial': pedido.historial.all(),
        'saldo': pedido.saldo_pendiente()
    })


@login_required
def perfil_cliente(request):
    """Ver y editar perfil del cliente"""
    if request.method == 'POST':
        user = request.user
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        
        if hasattr(user, 'telefono'):
            user.telefono = request.POST.get('telefono', '')
        
        user.save()
        messages.success(request, '✅ Perfil actualizado exitosamente')
        return redirect('perfil_cliente')
    
    return render(request, 'reportes/perfil_cliente.html', {
        'usuario': request.user
    })
    
    
# ==================== API PARA TASA DE CAMBIO ====================
@empleado_required
def api_tasa_actual(request):
    """API para obtener la tasa de cambio actual y precios de insumos"""
    tasa = TasaCambio.objects.first()
    tasa_valor = float(tasa.tasa) if tasa else 60.0
    
    insumos = Insumo.objects.filter(stock_actual__gt=0)
    insumos_data = []
    for insumo in insumos:
        insumos_data.append({
            'id': insumo.id,
            'nombre': insumo.nombre,
            'precio_usd': float(insumo.precio_usd),
            'precio_bs': float(insumo.precio_bs()),
            'stock': insumo.stock_actual,
        })
    
    return JsonResponse({
        'tasa': tasa_valor,
        'insumos': insumos_data,
        'fecha': tasa.fecha.strftime('%d/%m/%Y') if tasa else None
    })
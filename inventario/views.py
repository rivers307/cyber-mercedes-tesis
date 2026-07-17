import os
from django.conf import settings
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import models
from django.utils import timezone
from django.core.paginator import Paginator
from .models import (
    Insumo, Categoria, MovimientoInventario, AlertaStock,
    ActivoFijo, AuditoriaProgramada, InventarioTeorico, 
    HallazgoAuditoria, ConciliacionInventario, Auditoria
)
from sublimacion.models import Pedido
import json
from usuarios.decorators import empleado_required, admin_required
from reportes.utils import registrar_auditoria


# ========== DASHBOARD ==========
@empleado_required
def dashboard_inventario(request):
    insumos = Insumo.objects.all()
    activos = ActivoFijo.objects.all()
    
    total_insumos = insumos.count()
    stock_bajo = insumos.filter(stock_actual__lte=models.F('stock_minimo')).count()
    sin_stock = insumos.filter(stock_actual=0).count()
    valor_total = sum(insumo.stock_actual * float(insumo.precio_unitario) for insumo in insumos)
    
    total_activos = activos.count()
    activos_operativos = activos.filter(estado='operativo').count()
    activos_danados = activos.filter(estado='dañado').count()
    porcentaje_operatividad = (activos_operativos / total_activos * 100) if total_activos > 0 else 0
    
    alertas = AlertaStock.objects.filter(leida=False)[:10]
    ultimos_movimientos = MovimientoInventario.objects.all()[:15]
    auditorias_activas = AuditoriaProgramada.objects.filter(activa=True).count()
    
    context = {
        'total_insumos': total_insumos,
        'stock_bajo': stock_bajo,
        'sin_stock': sin_stock,
        'valor_total': valor_total,
        'alertas': alertas,
        'ultimos_movimientos': ultimos_movimientos,
        'insumos': insumos,
        'total_activos': total_activos,
        'activos_operativos': activos_operativos,
        'activos_danados': activos_danados,
        'porcentaje_operatividad': porcentaje_operatividad,
        'auditorias_activas': auditorias_activas,
    }
    return render(request, 'inventario/dashboard.html', context)


# ========== LISTA DE INSUMOS ==========
@empleado_required
def lista_insumos(request):
    insumos = Insumo.objects.all()
    categorias = Categoria.objects.all()
    return render(request, 'inventario/lista_insumos.html', {
        'insumos': insumos,
        'categorias': categorias
    })


# ========== REGISTRAR ENTRADA ==========
@empleado_required
@registrar_auditoria(
    accion='inventario_entrada',
    modulo='Inventario - Movimientos',
    obtener_descripcion=lambda r, **k: f"Registró entrada de {r.POST.get('cantidad', 0)} unidades de {r.POST.get('nombre', '')}"
)
def registrar_entrada(request):
    if request.method == 'POST':
        insumo_id = request.POST.get('insumo')
        cantidad = int(request.POST.get('cantidad'))
        motivo = request.POST.get('motivo', 'Compra de insumos')
        
        insumo = get_object_or_404(Insumo, id=insumo_id)
        stock_anterior = insumo.stock_actual
        stock_nuevo = stock_anterior + cantidad
        
        MovimientoInventario.objects.create(
            insumo=insumo,
            tipo='entrada',
            cantidad=cantidad,
            stock_anterior=stock_anterior,
            stock_nuevo=stock_nuevo,
            motivo=motivo,
            realizado_por=request.user
        )
        
        insumo.stock_actual = stock_nuevo
        insumo.save()
        
        messages.success(request, f'Se registró entrada de {cantidad} {insumo.unidad} de {insumo.nombre}')
        return redirect('lista_insumos')
    
    insumos = Insumo.objects.all()
    return render(request, 'inventario/registrar_entrada.html', {'insumos': insumos})


# ========== REGISTRAR SALIDA ==========
@empleado_required
@registrar_auditoria(
    accion='inventario_salida',
    modulo='Inventario - Movimientos',
    obtener_descripcion=lambda r, **k: f"Registró salida de {r.POST.get('cantidad', 0)} unidades de {r.POST.get('nombre', '')}"
)
def registrar_salida(request):
    if request.method == 'POST':
        insumo_id = request.POST.get('insumo')
        cantidad = int(request.POST.get('cantidad'))
        motivo = request.POST.get('motivo', 'Uso en producción')
        pedido_id = request.POST.get('pedido_relacionado')
        
        insumo = get_object_or_404(Insumo, id=insumo_id)
        
        if insumo.stock_actual < cantidad:
            messages.error(request, f'Stock insuficiente. Solo hay {insumo.stock_actual} {insumo.unidad}')
            return redirect('registrar_salida')
        
        stock_anterior = insumo.stock_actual
        stock_nuevo = stock_anterior - cantidad
        
        movimiento = MovimientoInventario.objects.create(
            insumo=insumo,
            tipo='salida',
            cantidad=cantidad,
            stock_anterior=stock_anterior,
            stock_nuevo=stock_nuevo,
            motivo=motivo,
            realizado_por=request.user
        )
        
        if pedido_id:
            pedido = get_object_or_404(Pedido, id=pedido_id)
            movimiento.pedido_relacionado = pedido
            movimiento.save()
        
        insumo.stock_actual = stock_nuevo
        insumo.save()
        
        if insumo.stock_actual <= insumo.stock_minimo:
            AlertaStock.objects.create(
                insumo=insumo,
                stock_actual=insumo.stock_actual,
                stock_minimo=insumo.stock_minimo
            )
            messages.warning(request, f'⚠️ Alerta: {insumo.nombre} está en stock mínimo')
        
        messages.success(request, f'Se registró salida de {cantidad} {insumo.unidad} de {insumo.nombre}')
        return redirect('lista_insumos')
    
    insumos = Insumo.objects.all()
    pedidos = Pedido.objects.filter(estado__in=['produccion', 'diseño'])[:20]
    return render(request, 'inventario/registrar_salida.html', {
        'insumos': insumos,
        'pedidos': pedidos
    })


# ========== CREAR INSUMO ==========
@admin_required
@registrar_auditoria(
    accion='crear',
    modulo='Inventario - Insumos',
    obtener_descripcion=lambda r, **k: f"Creó insumo {r.POST.get('nombre', '')} - Stock: {r.POST.get('stock_actual', 0)}"
)
def crear_insumo(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        categoria_id = request.POST.get('categoria')
        unidad = request.POST.get('unidad')
        stock_actual = int(request.POST.get('stock_actual', 0))
        stock_minimo = int(request.POST.get('stock_minimo', 5))
        precio_unitario = request.POST.get('precio_unitario', 0)
        proveedor = request.POST.get('proveedor', '')
        ubicacion = request.POST.get('ubicacion', '')
        
        insumo = Insumo.objects.create(
            nombre=nombre,
            categoria_id=categoria_id if categoria_id else None,
            unidad=unidad,
            stock_actual=stock_actual,
            stock_minimo=stock_minimo,
            precio_unitario=precio_unitario,
            proveedor=proveedor,
            ubicacion=ubicacion
        )
        
        insumo.codigo = f"INS-{insumo.id:04d}"
        insumo.save()
        
        messages.success(request, f'Insumo {nombre} creado exitosamente')
        return redirect('lista_insumos')
    
    categorias = Categoria.objects.all()
    return render(request, 'inventario/crear_insumo.html', {'categorias': categorias})


# ========== EDITAR INSUMO ==========
@admin_required
@registrar_auditoria(
    accion='editar',
    modulo='Inventario - Insumos',
    obtener_descripcion=lambda r, id, **k: f"Editó insumo #{id} - {r.POST.get('nombre', '')}"
)
def editar_insumo(request, id):
    insumo = get_object_or_404(Insumo, id=id)
    
    if request.method == 'POST':
        insumo.nombre = request.POST.get('nombre')
        insumo.categoria_id = request.POST.get('categoria')
        insumo.unidad = request.POST.get('unidad')
        insumo.stock_minimo = int(request.POST.get('stock_minimo', 5))
        insumo.precio_unitario = request.POST.get('precio_unitario', 0)
        insumo.proveedor = request.POST.get('proveedor', '')
        insumo.ubicacion = request.POST.get('ubicacion', '')
        insumo.notas = request.POST.get('notas', '')
        insumo.save()
        
        messages.success(request, f'Insumo {insumo.nombre} actualizado')
        return redirect('lista_insumos')
    
    categorias = Categoria.objects.all()
    return render(request, 'inventario/editar_insumo.html', {
        'insumo': insumo,
        'categorias': categorias
    })


# ========== ELIMINAR INSUMO ==========
@admin_required
@registrar_auditoria(
    accion='eliminar',
    modulo='Inventario - Insumos',
    obtener_descripcion=lambda r, id, **k: f"Eliminó insumo #{id} - {r.POST.get('nombre', '')}"
)
def eliminar_insumo(request, id):
    if request.method == 'POST':
        insumo = get_object_or_404(Insumo, id=id)
        nombre = insumo.nombre
        insumo.delete()
        messages.success(request, f'Insumo {nombre} eliminado')
        return redirect('lista_insumos')
    return redirect('lista_insumos')


# ========== HISTORIAL ==========
@empleado_required
def historial_movimientos(request):
    movimientos = MovimientoInventario.objects.all()
    
    tipo = request.GET.get('tipo')
    insumo_id = request.GET.get('insumo')
    
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    if insumo_id:
        movimientos = movimientos.filter(insumo_id=insumo_id)
    
    return render(request, 'inventario/historial.html', {
        'movimientos': movimientos,
        'insumos': Insumo.objects.all(),
        'tipos': MovimientoInventario.TIPOS_MOVIMIENTO
    })


# ========== MARCAR ALERTA ==========
@csrf_exempt
@empleado_required
def marcar_alerta_leida(request, id):
    if request.method == 'POST':
        alerta = get_object_or_404(AlertaStock, id=id)
        alerta.leida = True
        alerta.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})


# ========== CREAR CATEGORÍA ==========
@admin_required
@registrar_auditoria(
    accion='crear',
    modulo='Inventario - Categorías',
    obtener_descripcion=lambda r, **k: f"Creó categoría {r.POST.get('nombre', '')}"
)
def crear_categoria(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion', '')
        Categoria.objects.create(nombre=nombre, descripcion=descripcion)
        messages.success(request, f'Categoría "{nombre}" creada exitosamente')
        return redirect('lista_insumos')
    
    return render(request, 'inventario/crear_categoria.html')


# ========== GESTIÓN DE ACTIVOS ==========
@admin_required
@registrar_auditoria(
    accion='crear',
    modulo='Inventario - Activos Fijos',
    obtener_descripcion=lambda r, **k: f"Registró activo {r.POST.get('nombre', '')} - Código: {r.POST.get('codigo', '')}"
)
def gestion_activos(request):
    activos = ActivoFijo.objects.all()
    
    total = activos.count()
    operativos = activos.filter(estado='operativo').count()
    parciales = activos.filter(estado='parcial').count()
    danados = activos.filter(estado='dañado').count()
    obsoletos = activos.filter(estado='obsoleto').count()
    porcentaje_operatividad = (operativos / total * 100) if total > 0 else 0
    
    if request.method == 'POST':
        codigo = request.POST.get('codigo')
        nombre = request.POST.get('nombre')
        tipo = request.POST.get('tipo')
        estado = request.POST.get('estado')
        ubicacion = request.POST.get('ubicacion')
        marca = request.POST.get('marca', '')
        modelo = request.POST.get('modelo', '')
        valor_inicial = request.POST.get('valor_inicial', 0)
        fecha_adquisicion = request.POST.get('fecha_adquisicion') or None
        
        ActivoFijo.objects.create(
            codigo=codigo,
            nombre=nombre,
            tipo=tipo,
            estado=estado,
            ubicacion=ubicacion,
            marca=marca,
            modelo=modelo,
            valor_inicial=valor_inicial,
            fecha_adquisicion=fecha_adquisicion
        )
        messages.success(request, f'Activo {nombre} registrado')
        return redirect('gestion_activos')
    
    return render(request, 'inventario/activos.html', {
        'activos': activos,
        'total': total,
        'operativos': operativos,
        'parciales': parciales,
        'danados': danados,
        'obsoletos': obsoletos,
        'porcentaje_operatividad': porcentaje_operatividad
    })


# ========== AUDITORÍAS PROGRAMADAS ==========
@admin_required
def programar_auditoria(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        tipo = request.POST.get('tipo')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        
        auditoria = AuditoriaProgramada.objects.create(
            nombre=nombre,
            tipo=tipo,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            activa=True
        )
        
        for insumo in Insumo.objects.all():
            InventarioTeorico.objects.create(
                auditoria=auditoria,
                insumo=insumo,
                cantidad_teorica=insumo.stock_actual
            )
        
        for activo in ActivoFijo.objects.all():
            InventarioTeorico.objects.create(
                auditoria=auditoria,
                activo=activo,
                estado_teorico=activo.estado
            )
        
        messages.success(request, f'Auditoría "{nombre}" programada exitosamente')
        return redirect('lista_auditorias')
    
    return render(request, 'inventario/programar_auditoria.html')


@admin_required
def lista_auditorias(request):
    auditorias = AuditoriaProgramada.objects.all().order_by('-fecha_inicio')
    
    auditorias_activas = auditorias.filter(activa=True).count()
    auditorias_finalizadas = auditorias.filter(activa=False).count()
    
    from datetime import timedelta
    hoy = timezone.now().date()
    proximas_30_dias = auditorias.filter(
        activa=True,
        fecha_inicio__gte=hoy,
        fecha_inicio__lte=hoy + timedelta(days=30)
    ).count()
    
    proximas_auditorias = auditorias.filter(
        activa=True,
        fecha_inicio__gte=hoy
    ).order_by('fecha_inicio')[:5]
    
    return render(request, 'inventario/lista_auditorias.html', {
        'auditorias': auditorias,
        'auditorias_activas': auditorias_activas,
        'auditorias_finalizadas': auditorias_finalizadas,
        'proximas_30_dias': proximas_30_dias,
        'proximas_auditorias': proximas_auditorias,
    })


@admin_required
@registrar_auditoria(
    accion='realizar_auditoria',
    modulo='Inventario - Auditorías',
    obtener_descripcion=lambda r, id, **k: f"Realizó auditoría #{id} - {r.POST.get('nombre', '')}"
)
def realizar_auditoria(request, id):
    auditoria = get_object_or_404(AuditoriaProgramada, id=id)
    inventario_teorico = InventarioTeorico.objects.filter(auditoria=auditoria)
    
    if request.method == 'POST':
        for item in inventario_teorico.filter(insumo__isnull=False):
            cantidad_fisica = request.POST.get(f'cantidad_fisica_{item.id}')
            if cantidad_fisica:
                HallazgoAuditoria.objects.create(
                    auditoria=auditoria,
                    insumo=item.insumo,
                    cantidad_fisica=int(cantidad_fisica),
                    observaciones=request.POST.get(f'obs_{item.id}', ''),
                    registrado_por=request.user
                )
        
        for item in inventario_teorico.filter(activo__isnull=False):
            estado_fisico = request.POST.get(f'estado_fisico_{item.id}')
            if estado_fisico:
                HallazgoAuditoria.objects.create(
                    auditoria=auditoria,
                    activo=item.activo,
                    estado_fisico=estado_fisico,
                    observaciones=request.POST.get(f'obs_{item.id}', ''),
                    registrado_por=request.user
                )
        
        messages.success(request, 'Hallazgos registrados. Ahora puede proceder a la conciliación.')
        return redirect('conciliar_auditoria', auditoria.id)
    
    return render(request, 'inventario/realizar_auditoria.html', {
        'auditoria': auditoria,
        'inventario_teorico': inventario_teorico
    })


@admin_required
@registrar_auditoria(
    accion='conciliar_auditoria',
    modulo='Inventario - Auditorías',
    obtener_descripcion=lambda r, id, **k: f"Concilió auditoría #{id}"
)
def conciliar_auditoria(request, id):
    auditoria = get_object_or_404(AuditoriaProgramada, id=id)
    hallazgos = HallazgoAuditoria.objects.filter(auditoria=auditoria)
    
    if request.method == 'POST':
        for hallazgo in hallazgos:
            if hallazgo.insumo:
                teorico = InventarioTeorico.objects.get(auditoria=auditoria, insumo=hallazgo.insumo)
                variacion = teorico.cantidad_teorica - hallazgo.cantidad_fisica
                
                justificacion = request.POST.get(f'justificacion_{hallazgo.id}')
                detalle = request.POST.get(f'detalle_{hallazgo.id}', '')
                costo = request.POST.get(f'costo_{hallazgo.id}', 0)
                
                if variacion != 0:
                    ConciliacionInventario.objects.create(
                        auditoria=auditoria,
                        insumo=hallazgo.insumo,
                        cantidad_teorica=teorico.cantidad_teorica,
                        cantidad_fisica=hallazgo.cantidad_fisica,
                        variacion=variacion,
                        justificacion=justificacion,
                        detalle_justificacion=detalle,
                        costo_reposicion=costo,
                        conciliado_por=request.user
                    )
            
            elif hallazgo.activo:
                teorico = InventarioTeorico.objects.get(auditoria=auditoria, activo=hallazgo.activo)
                
                justificacion = request.POST.get(f'justificacion_{hallazgo.id}')
                detalle = request.POST.get(f'detalle_{hallazgo.id}', '')
                
                if teorico.estado_teorico != hallazgo.estado_fisico:
                    ConciliacionInventario.objects.create(
                        auditoria=auditoria,
                        activo=hallazgo.activo,
                        estado_teorico=teorico.estado_teorico,
                        estado_fisico=hallazgo.estado_fisico,
                        justificacion=justificacion,
                        detalle_justificacion=detalle,
                        conciliado_por=request.user
                    )
        
        auditoria.activa = False
        auditoria.save()
        messages.success(request, 'Auditoría conciliada exitosamente')
        return redirect('reporte_auditoria', auditoria.id)
    
    items_conciliacion = []
    total_items = 0
    faltantes = 0
    sobrantes = 0
    
    for hallazgo in hallazgos:
        if hallazgo.insumo:
            teorico = InventarioTeorico.objects.get(auditoria=auditoria, insumo=hallazgo.insumo)
            variacion = teorico.cantidad_teorica - hallazgo.cantidad_fisica
            total_items += 1
            if variacion > 0:
                faltantes += 1
            elif variacion < 0:
                sobrantes += 1
            items_conciliacion.append({
                'hallazgo': hallazgo,
                'teorico': teorico.cantidad_teorica,
                'variacion': variacion,
                'tipo': 'insumo'
            })
        else:
            teorico = InventarioTeorico.objects.get(auditoria=auditoria, activo=hallazgo.activo)
            total_items += 1
            if teorico.estado_teorico != hallazgo.estado_fisico:
                faltantes += 1
            items_conciliacion.append({
                'hallazgo': hallazgo,
                'teorico': teorico.estado_teorico,
                'tipo': 'activo'
            })
    
    return render(request, 'inventario/conciliar_auditoria.html', {
        'auditoria': auditoria,
        'items': items_conciliacion,
        'total_items': total_items,
        'faltantes': faltantes,
        'sobrantes': sobrantes,
        'justificaciones': ConciliacionInventario.JUSTIFICACIONES
    })


# ============================================================
# ========== REPORTE DE AUDITORÍA ==========
# ============================================================
@admin_required
def reporte_auditoria(request, id):
    from decimal import Decimal
    
    auditoria = get_object_or_404(AuditoriaProgramada, id=id)
    conciliaciones = ConciliacionInventario.objects.filter(auditoria=auditoria)
    
    monto_total_auditado = Decimal('0')
    ajustes_positivos = Decimal('0')
    ajustes_negativos = Decimal('0')
    
    for item in InventarioTeorico.objects.filter(auditoria=auditoria, insumo__isnull=False):
        if item.insumo and item.cantidad_teorica > 0:
            monto_total_auditado += Decimal(str(item.cantidad_teorica)) * item.insumo.precio_unitario
    
    for item in InventarioTeorico.objects.filter(auditoria=auditoria, activo__isnull=False):
        if item.activo and item.activo.valor_inicial:
            monto_total_auditado += item.activo.valor_inicial
    
    for conc in conciliaciones:
        if conc.insumo and conc.variacion:
            if conc.variacion > 0:
                ajustes_negativos += abs(conc.variacion) * conc.insumo.precio_unitario
            elif conc.variacion < 0:
                ajustes_positivos += abs(conc.variacion) * conc.insumo.precio_unitario
    
    costo_total_reposicion = sum(float(c.costo_reposicion) for c in conciliaciones if c.costo_reposicion)
    
    resumen_justificaciones = {}
    for justif in dict(ConciliacionInventario.JUSTIFICACIONES).keys():
        count = conciliaciones.filter(justificacion=justif).count()
        if count > 0:
            resumen_justificaciones[justif] = count
    
    codigo_reporte = f"RA-{auditoria.id:04d}"
    
    return render(request, 'inventario/reporte_auditoria.html', {
        'auditoria': auditoria,
        'conciliaciones': conciliaciones,
        'costo_total_reposicion': costo_total_reposicion,
        'resumen_justificaciones': resumen_justificaciones,
        'codigo_reporte': codigo_reporte,
        'monto_total_auditado': monto_total_auditado,
        'ajustes_positivos': ajustes_positivos,
        'ajustes_negativos': ajustes_negativos,
    })


# ============================================================
# ========== BITÁCORA DE AUDITORÍA (LÍNEA DE TIEMPO) ==========
# ============================================================
@admin_required
def ver_auditoria_inventario(request):
    if request.user.rol not in ['admin']:
        messages.error(request, 'No tienes permiso para ver la auditoría')
        return redirect('dashboard_inventario')
    
    auditorias = Auditoria.objects.all().order_by('-fecha')
    tipos_acciones = Auditoria.TIPO_ACCIONES
    
    # Filtros
    usuario = request.GET.get('usuario')
    accion = request.GET.get('accion')
    modulo = request.GET.get('modulo')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if usuario:
        auditorias = auditorias.filter(usuario__username__icontains=usuario)
    if accion:
        auditorias = auditorias.filter(accion=accion)
    if modulo:
        auditorias = auditorias.filter(modulo__icontains=modulo)
    if fecha_desde:
        try:
            fecha_desde_dt = timezone.datetime.strptime(fecha_desde, '%Y-%m-%d')
            auditorias = auditorias.filter(fecha__date__gte=fecha_desde_dt.date())
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_dt = timezone.datetime.strptime(fecha_hasta, '%Y-%m-%d')
            auditorias = auditorias.filter(fecha__date__lte=fecha_hasta_dt.date())
        except ValueError:
            pass
    
    paginator = Paginator(auditorias, 50)
    page = request.GET.get('page')
    auditorias_pag = paginator.get_page(page)
    
    # Fechas para el período
    if not fecha_desde and auditorias.exists():
        fecha_desde = auditorias.last().fecha.strftime('%Y-%m-%d')
    if not fecha_hasta and auditorias.exists():
        fecha_hasta = auditorias.first().fecha.strftime('%Y-%m-%d')
    
    return render(request, 'inventario/auditoria.html', {
        'auditorias': auditorias_pag,
        'tipos_acciones': tipos_acciones,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    })


# ========== EDITAR ACTIVO ==========
@admin_required
@registrar_auditoria(
    accion='editar',
    modulo='Inventario - Activos Fijos',
    obtener_descripcion=lambda r, id, **k: f"Editó activo #{id} - {r.POST.get('nombre', '')}"
)
def editar_activo(request, id):
    activo = get_object_or_404(ActivoFijo, id=id)
    
    if request.method == 'POST':
        activo.codigo = request.POST.get('codigo')
        activo.nombre = request.POST.get('nombre')
        activo.tipo = request.POST.get('tipo')
        activo.estado = request.POST.get('estado')
        activo.ubicacion = request.POST.get('ubicacion')
        activo.marca = request.POST.get('marca', '')
        activo.modelo = request.POST.get('modelo', '')
        activo.valor_inicial = request.POST.get('valor_inicial', 0)
        activo.fecha_adquisicion = request.POST.get('fecha_adquisicion') or None
        activo.notas = request.POST.get('notas', '')
        activo.save()
        
        messages.success(request, f'Activo {activo.nombre} actualizado exitosamente')
        return redirect('gestion_activos')
    
    return redirect('gestion_activos')


# ========== ELIMINAR ACTIVO ==========
@admin_required
@registrar_auditoria(
    accion='eliminar',
    modulo='Inventario - Activos Fijos',
    obtener_descripcion=lambda r, id, **k: f"Eliminó activo #{id} - {r.POST.get('nombre', '')}"
)
def eliminar_activo(request, id):
    activo = get_object_or_404(ActivoFijo, id=id)
    nombre = activo.nombre
    
    if request.method == 'POST':
        activo.delete()
        messages.success(request, f'Activo {nombre} eliminado')
        return redirect('gestion_activos')
    
    return redirect('gestion_activos')

# ============================================================
# ========== EXPORTAR REPORTE DE INVENTARIO A EXCEL ==========
# ============================================================
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
import os
from django.conf import settings

@admin_required
def exportar_reporte_excel(request):
    """Genera un reporte de inventario en Excel con formato profesional y logo"""
    
    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Inventario"
    
    # ========== 1. AGREGAR LOGO ==========
    # Ruta del logo en la carpeta static
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_cyber_comunal.png')
    
    try:
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            # Ajustar tamaño del logo (opcional)
            img.width = 120
            img.height = 80
            # Posicionar en la celda A1
            ws.add_image(img, 'A1')
            # Ajustar altura de fila para el logo
            ws.row_dimensions[1].height = 90
        else:
            # Si no encuentra el logo, escribir el nombre del local
            ws.merge_cells('A1:F1')
            cell = ws['A1']
            cell.value = "CYBER COMUNAL"
            cell.font = Font(name='Arial', size=20, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    except Exception as e:
        # Fallback: escribir el nombre
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = "CYBER COMUNAL"
        cell.font = Font(name='Arial', size=20, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Subtítulo
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "TECH & SUBLIMACIÓN"
    cell.font = Font(name='Arial', size=12, italic=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dirección y contacto
    ws.merge_cells('A3:F3')
    cell = ws['A3']
    cell.value = "Dirección: Av. Principal, C.C. Las Mercedes, Nivel 1, Local 4B, Ciudad de la Amistad"
    cell.font = Font(name='Arial', size=9)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A4:F4')
    cell = ws['A4']
    cell.value = "Teléfono: +58 412-5551234  |  Email: cyber.comunal.ve@gmail.com"
    cell.font = Font(name='Arial', size=9)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Línea separadora
    ws.merge_cells('A5:F5')
    cell = ws['A5']
    cell.border = Border(bottom=Side(style='medium', color='7C3AED'))
    
    # Título del reporte
    ws.merge_cells('A6:F6')
    cell = ws['A6']
    cell.value = "REPORTE DE AUDITORÍA DE INVENTARIO"
    cell.font = Font(name='Arial', size=16, bold=True, color='7C3AED')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Período y fecha de emisión
    hoy = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells('A7:F7')
    cell = ws['A7']
    cell.value = f"Período: MAYO - JUNIO 2026  |  Fecha de Emisión: {hoy}"
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Espacio
    ws.row_dimensions[8].height = 20
    
    # ========== 2. TABLA DE DATOS ==========
    headers = ['Item', 'Descripción', 'Especificaciones', 'Cantidad', 'Precio Unitario', 'Monto Total']
    col_widths = [15, 25, 30, 12, 18, 18]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=i, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    
    # Datos de insumos
    insumos = Insumo.objects.all().order_by('nombre')
    row_num = 10
    total_general = 0
    for insumo in insumos:
        cantidad = insumo.stock_actual
        precio = float(insumo.precio_unitario)
        monto = cantidad * precio
        total_general += monto
        
        ws.cell(row=row_num, column=1, value=insumo.nombre)
        ws.cell(row=row_num, column=2, value=insumo.categoria.nombre if insumo.categoria else "—")
        ws.cell(row=row_num, column=3, value=insumo.unidad)
        ws.cell(row=row_num, column=4, value=cantidad)
        ws.cell(row=row_num, column=5, value=precio)
        ws.cell(row=row_num, column=6, value=monto)
        
        for col in [5, 6]:
            cell = ws.cell(row=row_num, column=col)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
        
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
        row_num += 1
    
    # Fila de totales
    ws.merge_cells(f'A{row_num}:D{row_num}')
    cell = ws.cell(row=row_num, column=1, value="TOTAL")
    cell.font = Font(name='Arial', size=12, bold=True)
    cell.alignment = Alignment(horizontal='right')
    
    ws.cell(row=row_num, column=6, value=total_general).number_format = '#,##0.00'
    ws.cell(row=row_num, column=6).font = Font(bold=True)
    
    # Bordes a toda la tabla
    for r in range(9, row_num + 1):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # ========== 3. RESUMEN DE JUSTIFICACIONES ==========
    row_num += 3
    ws.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws.cell(row=row_num, column=1, value="JUSTIFICACIONES DE DIFERENCIAS")
    cell.font = Font(name='Arial', size=14, bold=True, color='7C3AED')
    cell.alignment = Alignment(horizontal='center')
    
    row_num += 1
    sub_headers = ['Item', 'Descripción', 'Especificaciones', 'Cantidad', 'Precio Unitario', 'Monto']
    for i, h in enumerate(sub_headers, 1):
        cell = ws.cell(row=row_num, column=i, value=h)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.fill = PatternFill(start_color='E8E4F0', end_color='E8E4F0', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:F{row_num}')
    ws.cell(row=row_num, column=1, value="No hay justificaciones registradas para este reporte general.").alignment = Alignment(horizontal='center')
    
    # ========== 4. TABLA DE REPOSICIÓN ==========
    row_num += 3
    ws.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws.cell(row=row_num, column=1, value="TABLA DE MONTO TOTAL DE REPOSICIÓN")
    cell.font = Font(name='Arial', size=14, bold=True, color='7C3AED')
    cell.alignment = Alignment(horizontal='center')
    
    row_num += 1
    for i, h in enumerate(sub_headers, 1):
        cell = ws.cell(row=row_num, column=i, value=h)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.fill = PatternFill(start_color='E8E4F0', end_color='E8E4F0', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:F{row_num}')
    ws.cell(row=row_num, column=1, value="No hay datos de reposición disponibles.").alignment = Alignment(horizontal='center')
    
    # ========== 5. FIRMAS ==========
    row_num += 3
    ws.merge_cells(f'A{row_num}:B{row_num}')
    cell = ws.cell(row=row_num, column=1, value="FIRMA AUDITOR RESPONSABLE")
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'C{row_num}:D{row_num}')
    cell = ws.cell(row=row_num, column=3, value="FIRMA COMISIÓN DE CONTRALORÍA")
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'E{row_num}:F{row_num}')
    cell = ws.cell(row=row_num, column=5, value="FIRMA DE RECEPCIÓN")
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    row_num += 1
    for col in [1, 3, 5]:
        ws.merge_cells(f'{get_column_letter(col)}{row_num}:{get_column_letter(col+1)}{row_num}')
        cell = ws.cell(row=row_num, column=col, value="_________________________")
        cell.alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="(admin / AD01)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=3, value="(Punto de Control)").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=5, value="(____ / ____ / ____)").alignment = Alignment(horizontal='center')
    
    # ========== 6. PIE DE PÁGINA ==========
    row_num += 3
    ws.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws.cell(row=row_num, column=1, value="Documento generado por el Sistema Inteligente de Gestión del Cyber Comunitario 'Ntra. Sra. de Mercedes'")
    cell.font = Font(name='Arial', size=9, italic=True, color='666666')
    cell.alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws.cell(row=row_num, column=1, value=f"Generado el: {hoy}")
    cell.font = Font(name='Arial', size=9, color='666666')
    cell.alignment = Alignment(horizontal='center')
    
    # ========== 7. CREAR RESPUESTA HTTP ==========
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_inventario_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    
    wb.save(response)
    return response


# ============================================================
# ========== EXPORTAR HISTORIAL DE MOVIMIENTOS A EXCEL ==========
# ============================================================
@empleado_required
def exportar_historial_excel(request):
    """Genera un reporte de movimientos de inventario en Excel con formato profesional"""
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # Obtener movimientos con filtros (igual que en la vista historial)
    movimientos = MovimientoInventario.objects.all().select_related('insumo', 'realizado_por', 'pedido_relacionado')
    
    # Aplicar filtros desde GET (si existen)
    tipo = request.GET.get('tipo')
    insumo_id = request.GET.get('insumo')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    if insumo_id:
        movimientos = movimientos.filter(insumo_id=insumo_id)
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            movimientos = movimientos.filter(fecha__date__gte=fecha_desde_dt.date())
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            movimientos = movimientos.filter(fecha__date__lte=fecha_hasta_dt.date())
        except ValueError:
            pass
    
    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Movimientos"
    
    # ========== 1. ENCABEZADO CON LOGO ==========
    # Logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_cyber_comunal.png')
    try:
        if os.path.exists(logo_path):
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            img.width = 120
            img.height = 80
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 90
        else:
            ws.merge_cells('A1:I1')
            ws['A1'] = "CYBER COMUNAL"
            ws['A1'].font = Font(name='Arial', size=20, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
    except Exception:
        ws.merge_cells('A1:I1')
        ws['A1'] = "CYBER COMUNAL"
        ws['A1'].font = Font(name='Arial', size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
    
    # Subtítulo
    ws.merge_cells('A2:I2')
    ws['A2'] = "TECH & SUBLIMACIÓN"
    ws['A2'].font = Font(name='Arial', size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Dirección
    ws.merge_cells('A3:I3')
    ws['A3'] = "Dirección: Av. Principal, C.C. Las Mercedes, Nivel 1, Local 4B, Ciudad de la Amistad"
    ws['A3'].font = Font(name='Arial', size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A4:I4')
    ws['A4'] = "Teléfono: +58 412-5551234  |  Email: cyber.comunal.ve@gmail.com"
    ws['A4'].font = Font(name='Arial', size=9)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    # Línea separadora
    ws.merge_cells('A5:I5')
    ws['A5'].border = Border(bottom=Side(style='medium', color='7C3AED'))
    
    # Título
    ws.merge_cells('A6:I6')
    ws['A6'] = "REPORTE DE MOVIMIENTOS DE INVENTARIO"
    ws['A6'].font = Font(name='Arial', size=16, bold=True, color='7C3AED')
    ws['A6'].alignment = Alignment(horizontal='center')
    
    hoy = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells('A7:I7')
    ws['A7'] = f"Fecha de Emisión: {hoy}"
    ws['A7'].font = Font(name='Arial', size=10, bold=True)
    ws['A7'].alignment = Alignment(horizontal='center')
    
    # ========== 2. TABLA DE DATOS ==========
    headers = ['Fecha/Hora', 'Insumo', 'Tipo', 'Cantidad', 'Stock Anterior', 'Stock Nuevo', 'Motivo', 'Usuario', 'Pedido Relacionado']
    col_widths = [20, 25, 18, 12, 15, 15, 30, 18, 20]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=i, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    
    # Datos
    row_num = 10
    for mov in movimientos:
        ws.cell(row=row_num, column=1, value=mov.fecha.strftime("%d/%m/%Y %H:%M:%S"))
        ws.cell(row=row_num, column=2, value=mov.insumo.nombre)
        ws.cell(row=row_num, column=3, value=mov.get_tipo_display())
        ws.cell(row=row_num, column=4, value=mov.cantidad)
        ws.cell(row=row_num, column=5, value=mov.stock_anterior)
        ws.cell(row=row_num, column=6, value=mov.stock_nuevo)
        ws.cell(row=row_num, column=7, value=mov.motivo or "—")
        ws.cell(row=row_num, column=8, value=mov.realizado_por.username)
        ws.cell(row=row_num, column=9, value=f"#{mov.pedido_relacionado.id}" if mov.pedido_relacionado else "—")
        
        # Alinear columnas específicas al centro
        for col in [4, 5, 6]:  # Cantidad, Stock Anterior, Stock Nuevo
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
        # Alinear Fecha/Hora al centro también
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        row_num += 1
    
    # Aplicar bordes a toda la tabla
    for r in range(9, row_num + 1):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # ========== 3. PIE DE PÁGINA ==========
    row_num += 2
    ws.merge_cells(f'A{row_num}:I{row_num}')
    ws.cell(row=row_num, column=1, value="Documento generado por el Sistema Inteligente de Gestión del Cyber Comunitario 'Ntra. Sra. de Mercedes'").font = Font(name='Arial', size=9, italic=True, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:I{row_num}')
    ws.cell(row=row_num, column=1, value=f"Generado el: {hoy}").font = Font(name='Arial', size=9, color='666666')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    # ========== 4. CREAR RESPUESTA HTTP ==========
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=historial_movimientos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    
    wb.save(response)
    return response